from openpyxl import load_workbook
my_text_1=""
my_str=""
my_str_1=""
my_text_2=""
read_exel=load_workbook("C://Users/dmoho/Desktop/TestCases.xlsx")
read=read_exel.active
for row in read.iter_rows(min_row=1,max_row=56,min_col=1,max_col=2,values_only=True):
    my_str=f'Eliya Sulimani - TC {row[0]} - {row[1]}'
    my_text_1 += f'{my_str}\n \n '

file_1=open('C:/Users/dmoho/Desktop/TestCases.txt',"w+")
file_1.write(my_text_1)
file_1.close()

# for row_1 in read.iter_rows(min_row=156,max_row=175,min_col=2,max_col=3,values_only=True):
#     my_str_1=f'Dima Farfel - HLD {row_1[0]} - {row_1[1]}'
#     my_text_2 += f'{my_str_1}\n \n'
# file_1=open('C:/Users/dmoho/Desktop/TestCases.txt',"a+")
# file_1.write(my_text_2)
# file_1.close()







