from openpyxl import Workbook,load_workbook
import random
a=random.randint(1000,9999)
b=random.randint(1000,9999)
workbook=Workbook()
sheet=workbook.active
sheet["A1"]="UserName"
sheet["A2"]="Tester_1"
sheet["A3"]="Tester_2"
sheet["B1"]="Password"
sheet["B2"]=a
sheet["B3"]=b
workbook.save('C:/Users/dmoho/Desktop/Login.xlsx')
load=load_workbook('C:/Users/dmoho/Desktop/Login.xlsx')
sheet=load.active
for cell in sheet.iter_rows(min_row=1,max_row=3,values_only=True):
    for data in cell:
        print(data)
for value in sheet.iter_cols(min_col=1,max_col=2,values_only=True):
    print(value)
