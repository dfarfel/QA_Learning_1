from openpyxl import load_workbook
load=load_workbook('C:/Users/dmoho/Desktop/Var.xlsx')
print(load.sheetnames)
sheet=load.active
a=sheet[5]
print(f'{sheet["A1"].value} {sheet["B1"].value} {a}')

