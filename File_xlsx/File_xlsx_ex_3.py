from openpyxl import load_workbook#Workbook
wb_read=load_workbook("C://Users/dmoho/Desktop/motfuck.xlsx")
# wb=Workbook()
# sheet=wb.create_sheet("Sheet1")
# sheet["A1"]="Hello kitty!"
# sheet["B1"]="ello kitty!"
# sheet["C1"]="llo kitty!"
# wb.save("C://Users/dmoho/Desktop/motfuck.xlsx")
abc=wb_read.active
c=abc["A1"]
a=abc.title
for row in abc.iter_rows(min_row=1,max_row=3,values_only=True):
    print(row)
    for value in row:
        print(value)