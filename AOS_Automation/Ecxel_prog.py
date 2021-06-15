from openpyxl import load_workbook


class Excel_AOS:
    def __init__(self):
        self.aos_table = load_workbook('C://Users//dmoho//Desktop//AOS.xlsx')
        self.read_table = self.aos_table.active

    # function that returns defined value of first product
    # str(Variable) - category,id,color,quantity
    # int(test_number) - from 1 to 10 include
    def product_1(self, variable, test_number):
        self.test_num = int(test_number)
        self.variable = str(variable)
        self.list_1 = []
        self.value = ''
        if self.variable == 'category':
            for row in self.read_table.iter_rows(min_row=2, max_row=2, min_col=3, max_col=12, values_only=True):
                for value in row:
                    self.list_1.append(value)
            self.value = self.list_1[self.test_num - 1]
        elif self.variable == 'id':
            for row in self.read_table.iter_rows(min_row=3, max_row=3, min_col=3, max_col=12, values_only=True):
                for value in row:
                    self.list_1.append(value)
            self.value = self.list_1[self.test_num - 1]
        elif self.variable == 'quantity':
            for row in self.read_table.iter_rows(min_row=4, max_row=4, min_col=3, max_col=12, values_only=True):
                for value in row:
                    self.list_1.append(value)
            self.value = self.list_1[self.test_num - 1]
        elif self.variable == 'color':
            for row in self.read_table.iter_rows(min_row=5, max_row=5, min_col=3, max_col=12, values_only=True):
                for value in row:
                    self.list_1.append(value)
            self.value = self.list_1[self.test_num - 1]
        self.list_1 = []
        return self.value

    # function that returns defined value of second product
    # str(Variable) - category,id,color,quantity
    # int(test_number) - from 1 to 10 include
    def product_2(self, variable, test_number):
        self.test_num_2 = int(test_number)
        self.variable_2 = str(variable)
        self.list_2 = []
        self.value_2 = ''
        if self.variable_2 == 'category':
            for row_2 in self.read_table.iter_rows(min_row=6, max_row=6, min_col=3, max_col=12, values_only=True):
                for value_2 in row_2:
                    self.list_2.append(value_2)
            self.value_2 = self.list_2[self.test_num_2 - 1]
        elif self.variable_2 == 'id':
            for row_2 in self.read_table.iter_rows(min_row=7, max_row=7, min_col=3, max_col=12, values_only=True):
                for value_2 in row_2:
                    self.list_2.append(value_2)
            self.value_2 = self.list_2[self.test_num_2 - 1]
        elif self.variable == 'quantity':
            for row_2 in self.read_table.iter_rows(min_row=8, max_row=8, min_col=3, max_col=12, values_only=True):
                for value_2 in row_2:
                    self.list_2.append(value_2)
            self.value_2 = self.list_2[self.test_num_2 - 1]
        elif self.variable == 'color':
            for row_2 in self.read_table.iter_rows(min_row=9, max_row=9, min_col=3, max_col=12, values_only=True):
                for value_2 in row_2:
                    self.list_2.append(value_2)
            self.value_2 = self.list_2[self.test_num_2 - 1]
        self.list_2 = []
        return self.value_2

    # function that returns defined value of third product
    # str(Variable) - category,id,color,quantity
    # int(test_number) - from 1 to 10 include
    def product_3(self, variable, test_number):
        self.test_num_3 = int(test_number)
        self.variable_3 = str(variable)
        self.list_3 = []
        self.value_3 = ''
        if self.variable_3 == 'category':
            for row_3 in self.read_table.iter_rows(min_row=10, max_row=10, min_col=3, max_col=12, values_only=True):
                for value_3 in row_3:
                    self.list_3.append(value_3)
            self.value_3 = self.list_3[self.test_num_3 - 1]
        elif self.variable_3 == 'id':
            for row_3 in self.read_table.iter_rows(min_row=11, max_row=11, min_col=3, max_col=12, values_only=True):
                for value_3 in row_3:
                    self.list_3.append(value_3)
            self.value_3 = self.list_3[self.test_num_3 - 1]
        elif self.variable_3 == 'quantity':
            for row_3 in self.read_table.iter_rows(min_row=12, max_row=12, min_col=3, max_col=12, values_only=True):
                for value_3 in row_3:
                    self.list_3.append(value_3)
            self.value_3 = self.list_3[self.test_num_3 - 1]
        elif self.variable_3 == 'color':
            for row_3 in self.read_table.iter_rows(min_row=13, max_row=13, min_col=3, max_col=12, values_only=True):
                for value_3 in row_3:
                    self.list_3.append(value_3)
            self.value_3 = self.list_3[self.test_num_3 - 1]
        self.list_3 = []
        return self.value_3

    # function that returns defined value of existing account
    # str(Variable) - username,password
    # int(test_number) - from 1 to 10 include
    def exist_account(self, variable, test_number):
        self.variable_4 = str(variable)
        self.test_num_4 = int(test_number)
        self.list_4 = []
        if self.variable_4 == 'username':
            for row_4 in self.read_table.iter_rows(min_row=14, max_row=14, min_col=3, max_col=12, values_only=True):
                for value_4 in row_4:
                    self.list_4.append(value_4)
            self.value_4 = self.list_4[self.test_num_4 - 1]
        elif self.variable_4 == 'password':
            for row_4 in self.read_table.iter_rows(min_row=15, max_row=15, min_col=3, max_col=12, values_only=True):
                for value_4 in row_4:
                    self.list_4.append(value_4)
            self.value_4 = self.list_4[self.test_num_4 - 1]
        self.list_4 = []
        return self.value_4

    # function that returns defined value of new account
    # str(Variable) - username,password,email
    # int(test_number) - from 1 to 10 include
    def new_account(self, variable, test_number):
        self.variable_5 = str(variable)
        self.test_num_5 = int(test_number)
        self.list_5 = []
        if self.variable_5 == 'username':
            for row_5 in self.read_table.iter_rows(min_row=16, max_row=16, min_col=3, max_col=12, values_only=True):
                for value_5 in row_5:
                    self.list_5.append(value_5)
            self.value_5 = self.list_5[self.test_num_5 - 1]
        elif self.variable_5 == 'password':
            for row_5 in self.read_table.iter_rows(min_row=17, max_row=17, min_col=3, max_col=12, values_only=True):
                for value_5 in row_5:
                    self.list_5.append(value_5)
            self.value_5 = self.list_5[self.test_num_5 - 1]
        elif self.variable_5 == 'email':
            for row_5 in self.read_table.iter_rows(min_row=18, max_row=18, min_col=3, max_col=12, values_only=True):
                for value_5 in row_5:
                    self.list_5.append(value_5)
            self.value_5 = self.list_5[self.test_num_5 - 1]
        self.list_5 = []
        return self.value_5

    # function that returns defined value of save pay account
    # str(Variable) - username,password
    # int(test_number) - from 1 to 10 include
    def save_pay(self, variable, test_number):
        self.variable_6 = str(variable)
        self.test_num_6 = int(test_number)
        self.list_6 = []
        if self.variable_6 == 'username':
            for row_6 in self.read_table.iter_rows(min_row=19, max_row=19, min_col=3, max_col=12, values_only=True):
                for value_6 in row_6:
                    self.list_6.append(value_6)
            self.value_6 = self.list_6[self.test_num_6 - 1]
        elif self.variable_6 == 'password':
            for row_6 in self.read_table.iter_rows(min_row=20, max_row=20, min_col=3, max_col=12, values_only=True):
                for value_6 in row_6:
                    self.list_6.append(value_6)
            self.value_6 = self.list_6[self.test_num_6 - 1]
        self.list_6 = []
        return self.value_6


    # function that returns defined value of master credit card
    # str(Variable) - number,cvv,month,year,name
    # int(test_number) - from 1 to 10 include
    def master_credit(self,variable,test_number):
        self.variable_7 = str(variable)
        self.test_num_7 = int(test_number)
        self.list_7 = []
        if self.variable_7 == 'number':
            for row_7 in self.read_table.iter_rows(min_row=21, max_row=21, min_col=3, max_col=12, values_only=True):
                for value_7 in row_7:
                    self.list_7.append(value_7)
            self.value_7 = self.list_7[self.test_num_7 - 1]
        elif self.variable_7 == 'cvv':
            for row_7 in self.read_table.iter_rows(min_row=22, max_row=22, min_col=3, max_col=12, values_only=True):
                for value_7 in row_7:
                    self.list_7.append(value_7)
            self.value_7 = self.list_7[self.test_num_7 - 1]
        elif self.variable_7 == 'month':
            for row_7 in self.read_table.iter_rows(min_row=23, max_row=23, min_col=3, max_col=12, values_only=True):
                for value_7 in row_7:
                    self.list_7.append(value_7)
            self.value_7 = self.list_7[self.test_num_7 - 1]
        elif self.variable_7 == 'year':
            for row_7 in self.read_table.iter_rows(min_row=24, max_row=24, min_col=3, max_col=12, values_only=True):
                for value_7 in row_7:
                    self.list_7.append(value_7)
            self.value_7 = self.list_7[self.test_num_7 - 1]
        elif self.variable_7 == 'name':
            for row_7 in self.read_table.iter_rows(min_row=25, max_row=25, min_col=3, max_col=12, values_only=True):
                for value_7 in row_7:
                    self.list_7.append(value_7)
            self.value_7 = self.list_7[self.test_num_7 - 1]
        self.list_7 = []
        return self.value_7